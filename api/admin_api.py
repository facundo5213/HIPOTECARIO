import getpass
import hashlib
import io
import json
import os
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from db import fetch_all, fetch_one, get_connection

router = APIRouter(prefix="/api/admin", tags=["Administración normativa"])
TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "plantillas" / "Plantilla_Mantenimiento_Normativo.xlsx"
MAX_FILE_SIZE = 10 * 1024 * 1024
ROLE_LEVEL = {"CONSULTOR": 0, "CARGADOR": 1, "APROBADOR": 2, "ADMINISTRADOR": 3}

SHEETS = {
    "Productos": {
        "type": "PRODUCTO", "key": ("Codigo",),
        "headers": ("Accion", "Codigo", "Activo"),
    },
    "Equivalencias": {
        "type": "EQUIVALENCIA", "key": ("Codigo_Desembolso", "Reglamentacion", "Fecha_Desde"),
        "headers": ("Accion", "Codigo_Desembolso", "Codigo_Consolidacion", "Reglamentacion", "Fecha_Desde", "Fecha_Hasta"),
    },
    "Condiciones": {
        "type": "CONDICION", "key": ("Clave_Mantenimiento",),
        "headers": (
            "Accion", "Clave_Mantenimiento", "Reglamentacion", "Ambito_Geografico", "Destino", "Tipo_Usuario",
            "Modalidad_Convenio", "Fecha_Desde", "Fecha_Hasta", "UVA_Desde", "UVA_Desde_Inclusive",
            "UVA_Hasta", "UVA_Hasta_Inclusive", "Tasa_Aplicable_Pct", "Permite_Topeo", "Cobra_Prima_Topeo",
            "Codigo_Producto", "Codigo_Desembolso", "Codigo_Consolidacion", "Referencia_Fuente", "Observaciones",
        ),
    },
    "Organismos": {
        "type": "ORGANISMO", "key": ("Codigo",),
        "headers": ("Accion", "Codigo", "Nombre", "CUIT", "Activo"),
    },
    "Organismo_Tasas": {
        "type": "ORGANISMO_TASA", "key": ("Codigo_Organismo", "Tipo_Vivienda", "Vigencia_Desde"),
        "headers": (
            "Accion", "Codigo_Organismo", "Tipo_Vivienda", "Tasa_Pct", "Grupo_Pauta", "Vigencia_Desde",
            "Vigencia_Hasta", "Adicional_Topeo_Pct", "Aplica_Circular_3214", "Observaciones",
        ),
    },
}


class ManualChange(BaseModel):
    tipo_registro: str = Field(pattern=r"^(PRODUCTO|EQUIVALENCIA|CONDICION|ORGANISMO|ORGANISMO_TASA)$")
    accion: str = Field(pattern=r"^(GUARDAR|BAJA)$")
    datos: dict
    observaciones: str | None = Field(default=None, max_length=1000)


class Decision(BaseModel):
    observaciones: str | None = Field(default=None, max_length=1000)


def windows_user(request: Request) -> str:
    if os.getenv("TRUST_WINDOWS_AUTH_HEADER", "false").lower() == "true":
        forwarded = request.headers.get("X-Remote-User")
        if forwarded:
            return forwarded.strip()
    domain = os.getenv("USERDOMAIN", "").strip()
    username = os.getenv("USERNAME", "").strip() or getpass.getuser()
    return f"{domain}\\{username}" if domain and "\\" not in username else username


def current_user(request: Request) -> dict:
    identity = windows_user(request)
    user = fetch_one("""
        SELECT Usuario_Windows AS usuario,COALESCE(Nombre_Mostrar,Usuario_Windows) AS nombre,Rol AS rol
        FROM dbo.Usuarios_Aplicacion WHERE Activo=1 AND UPPER(Usuario_Windows)=UPPER(?)
    """, (identity,))
    if not user:
        raise HTTPException(status_code=403, detail=f"El usuario Windows {identity} no tiene permisos en Administración")
    return user


def require_role(minimum: str):
    def dependency(user: dict = Depends(current_user)) -> dict:
        if ROLE_LEVEL[user["rol"]] < ROLE_LEVEL[minimum]:
            raise HTTPException(status_code=403, detail=f"Esta operación requiere rol {minimum} o superior")
        return user
    return dependency


def json_default(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    raise TypeError(type(value).__name__)


def dump_json(value) -> str:
    return json.dumps(value, ensure_ascii=False, default=json_default, sort_keys=True)


def empty(value) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def text_value(value, upper=False):
    if empty(value):
        return None
    result = str(value).strip()
    return result.upper() if upper else result


def date_value(value, field, errors):
    if empty(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            pass
    errors.append((field, "FECHA_INVALIDA", "Debe utilizar una fecha válida (DD/MM/AAAA)."))
    return None


def decimal_value(value, field, errors):
    if empty(value):
        return None
    try:
        return Decimal(str(value).strip().replace("%", "").replace(",", "."))
    except InvalidOperation:
        errors.append((field, "NUMERO_INVALIDO", "Debe ser un número válido."))
        return None


def bool_value(value, field, errors, nullable=True):
    if empty(value):
        return None if nullable else False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in (0, 1):
        return bool(value)
    normalized = str(value).strip().upper()
    if normalized in {"SI", "SÍ", "S", "TRUE", "1", "ACTIVO"}:
        return True
    if normalized in {"NO", "N", "FALSE", "0", "INACTIVO"}:
        return False
    errors.append((field, "BOOLEANO_INVALIDO", "Debe indicar SI o NO."))
    return None


def references() -> dict:
    return {
        "regulations": {row["codigo"] for row in fetch_all("SELECT Codigo AS codigo FROM dbo.Reglamentaciones")},
        "areas": {row["codigo"] for row in fetch_all("SELECT Codigo AS codigo FROM dbo.Ambitos_Geograficos")},
        "destinations": {row["codigo"] for row in fetch_all("SELECT Codigo AS codigo FROM dbo.Destinos")},
        "users": {row["codigo"] for row in fetch_all("SELECT Codigo AS codigo FROM dbo.Tipos_Usuario")},
        "modalities": {(row["reglamentacion"], row["codigo"]) for row in fetch_all("""
            SELECT r.Codigo AS reglamentacion,mc.Codigo AS codigo FROM dbo.Modalidades_Convenio mc
            JOIN dbo.Reglamentaciones r ON r.Reglamentacion_ID=mc.Reglamentacion_ID
        """)},
        "products": {row["codigo"] for row in fetch_all("SELECT Codigo AS codigo FROM dbo.Productos")},
        "organisms": {row["codigo"] for row in fetch_all("SELECT Codigo AS codigo FROM dbo.Organismos")},
    }


def normalize_record(record_type: str, raw: dict, refs: dict) -> tuple[dict, list]:
    errors = []
    data = dict(raw)
    action = text_value(data.pop("Accion", "GUARDAR"), True) or "GUARDAR"
    if action not in {"GUARDAR", "BAJA"}:
        errors.append(("Accion", "ACCION_INVALIDA", "La acción debe ser GUARDAR o BAJA."))

    if record_type == "PRODUCTO":
        data = {"Codigo": text_value(data.get("Codigo"), True), "Activo": bool_value(data.get("Activo"), "Activo", errors)}
        if not data["Codigo"]: errors.append(("Codigo", "REQUERIDO", "El código es obligatorio."))
    elif record_type == "ORGANISMO":
        data = {"Codigo": text_value(data.get("Codigo"), True), "Nombre": text_value(data.get("Nombre")), "CUIT": text_value(data.get("CUIT")), "Activo": bool_value(data.get("Activo"), "Activo", errors)}
        for field in (("Codigo",) if action == "BAJA" else ("Codigo", "Nombre", "CUIT")):
            if not data[field]: errors.append((field, "REQUERIDO", f"{field} es obligatorio."))
        if data["CUIT"] and (len(data["CUIT"]) != 11 or not data["CUIT"].isdigit()): errors.append(("CUIT", "CUIT_INVALIDO", "El CUIT debe contener exactamente 11 dígitos."))
    elif record_type == "EQUIVALENCIA":
        data = {
            "Codigo_Desembolso": text_value(data.get("Codigo_Desembolso"), True), "Codigo_Consolidacion": text_value(data.get("Codigo_Consolidacion"), True),
            "Reglamentacion": text_value(data.get("Reglamentacion"), True), "Fecha_Desde": date_value(data.get("Fecha_Desde"), "Fecha_Desde", errors),
            "Fecha_Hasta": date_value(data.get("Fecha_Hasta"), "Fecha_Hasta", errors),
        }
        for field in (("Codigo_Desembolso", "Reglamentacion") if action == "BAJA" else ("Codigo_Desembolso", "Reglamentacion", "Fecha_Desde")):
            if not data[field]: errors.append((field, "REQUERIDO", f"{field} es obligatorio."))
        if action == "GUARDAR" and not data["Codigo_Consolidacion"]: errors.append(("Codigo_Consolidacion", "REQUERIDO", "El producto de consolidación es obligatorio."))
        if data["Reglamentacion"] and data["Reglamentacion"] not in refs["regulations"]: errors.append(("Reglamentacion", "MAESTRO_INEXISTENTE", "La reglamentación no existe."))
    elif record_type == "ORGANISMO_TASA":
        data = {
            "Codigo_Organismo": text_value(data.get("Codigo_Organismo"), True), "Tipo_Vivienda": text_value(data.get("Tipo_Vivienda"), True),
            "Tasa_Pct": decimal_value(data.get("Tasa_Pct"), "Tasa_Pct", errors), "Grupo_Pauta": text_value(data.get("Grupo_Pauta"), True),
            "Vigencia_Desde": date_value(data.get("Vigencia_Desde"), "Vigencia_Desde", errors), "Vigencia_Hasta": date_value(data.get("Vigencia_Hasta"), "Vigencia_Hasta", errors),
            "Adicional_Topeo_Pct": decimal_value(data.get("Adicional_Topeo_Pct"), "Adicional_Topeo_Pct", errors),
            "Aplica_Circular_3214": bool_value(data.get("Aplica_Circular_3214"), "Aplica_Circular_3214", errors), "Observaciones": text_value(data.get("Observaciones")),
        }
        required_org_rate = ("Codigo_Organismo", "Tipo_Vivienda", "Vigencia_Hasta") if action == "BAJA" else ("Codigo_Organismo", "Tipo_Vivienda", "Vigencia_Desde", "Vigencia_Hasta")
        for field in required_org_rate:
            if not data[field]: errors.append((field, "REQUERIDO", f"{field} es obligatorio."))
        if action == "GUARDAR" and data["Tasa_Pct"] is None: errors.append(("Tasa_Pct", "REQUERIDO", "La tasa es obligatoria."))
        if data["Tipo_Vivienda"] not in {"UNICA", "SEGUNDA"}: errors.append(("Tipo_Vivienda", "VALOR_INVALIDO", "Debe ser UNICA o SEGUNDA."))
        if action == "GUARDAR" and data["Grupo_Pauta"] not in {"PRE_3214", "POST_3214"}: errors.append(("Grupo_Pauta", "VALOR_INVALIDO", "Debe ser PRE_3214 o POST_3214."))
        if action == "GUARDAR" and data["Tipo_Vivienda"] == "UNICA" and data["Adicional_Topeo_Pct"] is None: errors.append(("Adicional_Topeo_Pct", "REQUERIDO", "Primera vivienda requiere adicional de topeo."))
    elif record_type == "CONDICION":
        fields_upper = ("Clave_Mantenimiento", "Reglamentacion", "Ambito_Geografico", "Destino", "Tipo_Usuario", "Modalidad_Convenio", "Codigo_Producto", "Codigo_Desembolso", "Codigo_Consolidacion")
        data = {field: text_value(raw.get(field), True) for field in fields_upper}
        data.update({
            "Fecha_Desde": date_value(raw.get("Fecha_Desde"), "Fecha_Desde", errors), "Fecha_Hasta": date_value(raw.get("Fecha_Hasta"), "Fecha_Hasta", errors),
            "UVA_Desde": decimal_value(raw.get("UVA_Desde"), "UVA_Desde", errors), "UVA_Desde_Inclusive": bool_value(raw.get("UVA_Desde_Inclusive"), "UVA_Desde_Inclusive", errors),
            "UVA_Hasta": decimal_value(raw.get("UVA_Hasta"), "UVA_Hasta", errors), "UVA_Hasta_Inclusive": bool_value(raw.get("UVA_Hasta_Inclusive"), "UVA_Hasta_Inclusive", errors),
            "Tasa_Aplicable_Pct": decimal_value(raw.get("Tasa_Aplicable_Pct"), "Tasa_Aplicable_Pct", errors),
            "Permite_Topeo": bool_value(raw.get("Permite_Topeo"), "Permite_Topeo", errors), "Cobra_Prima_Topeo": bool_value(raw.get("Cobra_Prima_Topeo"), "Cobra_Prima_Topeo", errors),
            "Referencia_Fuente": text_value(raw.get("Referencia_Fuente")), "Observaciones": text_value(raw.get("Observaciones")),
        })
        required = ("Clave_Mantenimiento",) if action == "BAJA" else ("Clave_Mantenimiento", "Reglamentacion", "Ambito_Geografico", "Destino", "Tipo_Usuario", "Tasa_Aplicable_Pct", "Referencia_Fuente")
        for field in required:
            if data.get(field) is None: errors.append((field, "REQUERIDO", f"{field} es obligatorio."))
        checks = (("Reglamentacion", "regulations"), ("Ambito_Geografico", "areas"), ("Destino", "destinations"), ("Tipo_Usuario", "users"))
        for field, ref_name in checks:
            if data.get(field) and data[field] not in refs[ref_name]: errors.append((field, "MAESTRO_INEXISTENTE", f"{field} no existe en la base."))
        if data.get("Cobra_Prima_Topeo") and not data.get("Permite_Topeo"): errors.append(("Cobra_Prima_Topeo", "REGLA_NEGOCIO", "No puede cobrar prima si no permite topeo."))
        if action == "GUARDAR" and not any(data.get(field) for field in ("Codigo_Producto", "Codigo_Desembolso")): errors.append(("Codigo_Producto", "PRODUCTO_REQUERIDO", "Debe indicar producto único o producto de desembolso."))
    else:
        errors.append((None, "TIPO_INVALIDO", "Tipo de registro desconocido."))

    start = data.get("Fecha_Desde") or data.get("Vigencia_Desde")
    end = data.get("Fecha_Hasta") or data.get("Vigencia_Hasta")
    if start and end and end < start: errors.append(("Fecha_Hasta", "RANGO_INVALIDO", "La fecha hasta no puede ser anterior a la fecha desde."))
    return {"action": action, "data": data}, errors


def current_snapshot(record_type: str, data: dict) -> dict | None:
    if record_type == "PRODUCTO":
        return fetch_one("SELECT Codigo,Activo FROM dbo.Productos WHERE Codigo=?", (data["Codigo"],))
    if record_type == "ORGANISMO":
        return fetch_one("SELECT Codigo,Nombre,CUIT,Activo FROM dbo.Organismos WHERE Codigo=?", (data["Codigo"],))
    if record_type == "EQUIVALENCIA":
        return fetch_one("""SELECT TOP(1) pd.Codigo AS Codigo_Desembolso,pc.Codigo AS Codigo_Consolidacion,r.Codigo AS Reglamentacion,pe.Fecha_Desde,pe.Fecha_Hasta,pe.Activa
            FROM dbo.Producto_Equivalencias pe JOIN dbo.Productos pd ON pd.Producto_ID=pe.Producto_Desembolso_ID
            JOIN dbo.Productos pc ON pc.Producto_ID=pe.Producto_Consolidacion_ID JOIN dbo.Reglamentaciones r ON r.Reglamentacion_ID=pe.Reglamentacion_ID
            WHERE pd.Codigo=? AND r.Codigo=? ORDER BY pe.Fecha_Desde DESC""", (data["Codigo_Desembolso"], data["Reglamentacion"]))
    if record_type == "CONDICION":
        return fetch_one("""SELECT TOP(1) cr.Clave_Mantenimiento,cr.Codigo_Condicion,r.Codigo AS Reglamentacion,ag.Codigo AS Ambito_Geografico,d.Codigo AS Destino,tu.Codigo AS Tipo_Usuario,
            mc.Codigo AS Modalidad_Convenio,cr.Fecha_Desde,cr.Fecha_Hasta,cr.UVA_Desde,cr.UVA_Desde_Inclusive,cr.UVA_Hasta,cr.UVA_Hasta_Inclusive,
            cr.Tasa_Aplicable_Pct,cr.Permite_Topeo,cr.Cobra_Prima_Topeo,
            (SELECT p.Codigo FROM dbo.Condicion_Productos cp JOIN dbo.Productos p ON p.Producto_ID=cp.Producto_ID WHERE cp.Condicion_ID=cr.Condicion_ID AND cp.Rol_Producto='UNICO') AS Codigo_Producto,
            (SELECT p.Codigo FROM dbo.Condicion_Productos cp JOIN dbo.Productos p ON p.Producto_ID=cp.Producto_ID WHERE cp.Condicion_ID=cr.Condicion_ID AND cp.Rol_Producto='DESEMBOLSO') AS Codigo_Desembolso,
            (SELECT p.Codigo FROM dbo.Condicion_Productos cp JOIN dbo.Productos p ON p.Producto_ID=cp.Producto_ID WHERE cp.Condicion_ID=cr.Condicion_ID AND cp.Rol_Producto='CONSOLIDACION') AS Codigo_Consolidacion,
            cr.Referencia_Fuente,cr.Observaciones,cr.Activa
            FROM dbo.Condiciones_Reglamentacion cr JOIN dbo.Reglamentaciones r ON r.Reglamentacion_ID=cr.Reglamentacion_ID
            JOIN dbo.Ambitos_Geograficos ag ON ag.Ambito_Geografico_ID=cr.Ambito_Geografico_ID JOIN dbo.Destinos d ON d.Destino_ID=cr.Destino_ID
            JOIN dbo.Tipos_Usuario tu ON tu.Tipo_Usuario_ID=cr.Tipo_Usuario_ID LEFT JOIN dbo.Modalidades_Convenio mc ON mc.Modalidad_Convenio_ID=cr.Modalidad_Convenio_ID
            WHERE cr.Clave_Mantenimiento=? ORDER BY ISNULL(cr.Fecha_Desde,'19000101') DESC,cr.Condicion_ID DESC""", (data["Clave_Mantenimiento"],))
    if record_type == "ORGANISMO_TASA":
        return fetch_one("""SELECT TOP(1) o.Codigo AS Codigo_Organismo,oc.Tipo_Vivienda,mc.Tasa_Referencia_Pct AS Tasa_Pct,oc.Grupo_Pauta,
            oc.Vigencia_Desde,oc.Vigencia_Hasta,oc.Adicional_Topeo_Pct,oc.Aplica_Circular_3214,oc.Observaciones,oc.Activa
            FROM dbo.Organismo_Condiciones_802_01 oc JOIN dbo.Organismos o ON o.Organismo_ID=oc.Organismo_ID
            JOIN dbo.Modalidades_Convenio mc ON mc.Modalidad_Convenio_ID=oc.Modalidad_Convenio_ID
            WHERE o.Codigo=? AND oc.Tipo_Vivienda=? ORDER BY oc.Vigencia_Desde DESC""", (data["Codigo_Organismo"], data["Tipo_Vivienda"]))
    return None


def record_key(record_type: str, data: dict) -> str:
    fields = {config["type"]: config["key"] for config in SHEETS.values()}[record_type]
    return "|".join(str(data.get(field) or "") for field in fields)


def detect_action(requested: str, previous: dict | None, data: dict) -> str:
    if requested == "BAJA": return "BAJA"
    if previous is None: return "ALTA"
    comparable_previous = dump_json({key: previous.get(key) for key in data})
    comparable_new = dump_json(data)
    return "SIN_CAMBIOS" if comparable_previous == comparable_new else "MODIFICACION"


def persist_load(load_type: str, filename: str | None, content: bytes | None, observations: str | None, user: dict, records: list, errors: list) -> int:
    digest = hashlib.sha256(content).hexdigest() if content else None
    if digest and fetch_one("SELECT Carga_ID FROM dbo.Cargas_Normativas WHERE Hash_Archivo=?", (digest,)):
        raise HTTPException(status_code=409, detail="Este mismo archivo ya fue cargado anteriormente")
    valid_count = sum(1 for item in records if not item["errors"])
    error_count = sum(1 for item in errors if item["severity"] == "ERROR") + sum(len(item["errors"]) for item in records)
    state = "CON_ERRORES" if error_count else "VALIDADA"
    with get_connection() as connection:
        cursor = connection.cursor()
        cursor.execute("""INSERT dbo.Cargas_Normativas
            (Tipo_Carga,Nombre_Archivo,Hash_Archivo,Archivo_Original,Estado,Usuario_Carga,Observaciones,Total_Registros,Total_Errores)
            OUTPUT INSERTED.Carga_ID VALUES (?,?,?,?,?,?,?,?,?)""",
            (load_type, filename, digest, content, state, user["usuario"], observations, len(records), error_count))
        load_id = int(cursor.fetchone()[0])
        actions = {"ALTA": 0, "MODIFICACION": 0, "BAJA": 0}
        for item in records:
            normalized, item_errors = item["normalized"], item["errors"]
            data = normalized["data"]
            previous = current_snapshot(item["type"], data) if not item_errors else None
            detected = detect_action(normalized["action"], previous, data) if not item_errors else None
            if detected in actions: actions[detected] += 1
            cursor.execute("""INSERT dbo.Carga_Normativa_Detalles
                (Carga_ID,Hoja,Numero_Fila,Tipo_Registro,Clave_Registro,Accion_Solicitada,Accion_Detectada,Datos_JSON,Datos_Anteriores_JSON,Estado)
                VALUES (?,?,?,?,?,?,?,?,?,?)""", (load_id, item["sheet"], item["row"], item["type"], record_key(item["type"], data), normalized["action"], detected, dump_json(data), dump_json(previous) if previous else None, "ERROR" if item_errors else "VALIDO"))
            for field, code, message in item_errors:
                cursor.execute("INSERT dbo.Carga_Normativa_Errores (Carga_ID,Hoja,Numero_Fila,Campo,Severidad,Codigo_Error,Mensaje) VALUES (?,?,?,?,?,?,?)", (load_id, item["sheet"], item["row"], field, "ERROR", code, message))
        for issue in errors:
            cursor.execute("INSERT dbo.Carga_Normativa_Errores (Carga_ID,Hoja,Numero_Fila,Campo,Severidad,Codigo_Error,Mensaje) VALUES (?,?,?,?,?,?,?)", (load_id, issue.get("sheet"), issue.get("row"), issue.get("field"), issue["severity"], issue["code"], issue["message"]))
        cursor.execute("UPDATE dbo.Cargas_Normativas SET Total_Altas=?,Total_Modificaciones=?,Total_Bajas=? WHERE Carga_ID=?", (actions["ALTA"], actions["MODIFICACION"], actions["BAJA"], load_id))
        connection.commit()
    return load_id


def parse_workbook(content: bytes) -> tuple[list, list]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="El archivo no es un Excel .xlsx válido") from exc
    refs = references()
    records, general_errors, seen = [], [], set()
    for sheet_name, config in SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            general_errors.append({"sheet": sheet_name, "severity": "ERROR", "code": "HOJA_FALTANTE", "message": f"Falta la hoja obligatoria {sheet_name}."})
            continue
        sheet = workbook[sheet_name]
        if sheet.max_row > 5001:
            general_errors.append({"sheet": sheet_name, "severity": "ERROR", "code": "DEMASIADAS_FILAS", "message": "La hoja supera el máximo de 5.000 filas por carga."})
            continue
        rows = sheet.iter_rows(values_only=True)
        try: headers = tuple(text_value(value) for value in next(rows))
        except StopIteration: headers = ()
        if headers[:len(config["headers"])] != config["headers"]:
            general_errors.append({"sheet": sheet_name, "row": 1, "severity": "ERROR", "code": "ENCABEZADOS_INVALIDOS", "message": "Los encabezados fueron modificados. Descargue una plantilla nueva."})
            continue
        for row_number, values in enumerate(rows, start=2):
            if all(empty(value) for value in values[:len(config["headers"])]): continue
            raw = dict(zip(config["headers"], values))
            normalized, item_errors = normalize_record(config["type"], raw, refs)
            key = (config["type"], record_key(config["type"], normalized["data"]))
            if key in seen: item_errors.append((None, "DUPLICADO_ARCHIVO", "La misma clave aparece más de una vez en el archivo."))
            seen.add(key)
            records.append({"sheet": sheet_name, "row": row_number, "type": config["type"], "normalized": normalized, "errors": item_errors})
    staged_products = refs["products"] | {item["normalized"]["data"].get("Codigo") for item in records if item["type"] == "PRODUCTO" and item["normalized"]["action"] == "GUARDAR"}
    staged_organisms = refs["organisms"] | {item["normalized"]["data"].get("Codigo") for item in records if item["type"] == "ORGANISMO" and item["normalized"]["action"] == "GUARDAR"}
    for item in records:
        data, item_errors = item["normalized"]["data"], item["errors"]
        if item["type"] == "EQUIVALENCIA":
            for field in ("Codigo_Desembolso", "Codigo_Consolidacion"):
                if data.get(field) and data[field] not in staged_products: item_errors.append((field, "PRODUCTO_INEXISTENTE", f"El producto {data[field]} no existe ni está dado de alta en este archivo."))
        elif item["type"] == "CONDICION":
            for field in ("Codigo_Producto", "Codigo_Desembolso", "Codigo_Consolidacion"):
                if data.get(field) and data[field] not in staged_products: item_errors.append((field, "PRODUCTO_INEXISTENTE", f"El producto {data[field]} no existe ni está dado de alta en este archivo."))
            modality = data.get("Modalidad_Convenio")
            if modality and (data.get("Reglamentacion"), modality) not in refs["modalities"]: item_errors.append(("Modalidad_Convenio", "MODALIDAD_INVALIDA", "La modalidad no corresponde a la reglamentación."))
        elif item["type"] == "ORGANISMO_TASA":
            if data.get("Codigo_Organismo") not in staged_organisms: item_errors.append(("Codigo_Organismo", "ORGANISMO_INEXISTENTE", "El organismo no existe ni está dado de alta en este archivo."))
            if data.get("Tasa_Pct") is not None:
                modality = f"TASA_{int(data['Tasa_Pct'] * 100):03d}"
                if ("802_01", modality) not in refs["modalities"]: item_errors.append(("Tasa_Pct", "TASA_SIN_CUADRO", "No existe un cuadro/modalidad 802_01 para esta tasa."))
    return records, general_errors


@router.get("/me")
def me(user: dict = Depends(current_user)):
    return user


@router.get("/plantilla")
def template(_: dict = Depends(require_role("CARGADOR"))):
    if not TEMPLATE_PATH.exists(): raise HTTPException(status_code=404, detail="La plantilla no está instalada en el servidor")
    return FileResponse(TEMPLATE_PATH, filename=TEMPLATE_PATH.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/cargas")
async def upload_load(file: UploadFile = File(...), observaciones: str | None = Form(default=None), user: dict = Depends(require_role("CARGADOR"))):
    if not file.filename or not file.filename.lower().endswith(".xlsx"): raise HTTPException(status_code=422, detail="Debe subir la plantilla en formato .xlsx")
    content = await file.read(MAX_FILE_SIZE + 1)
    if len(content) > MAX_FILE_SIZE: raise HTTPException(status_code=413, detail="El archivo supera el máximo de 10 MB")
    records, errors = parse_workbook(content)
    if not records and not errors: raise HTTPException(status_code=422, detail="La plantilla no contiene registros")
    load_id = persist_load("EXCEL", file.filename, content, observations=observaciones, user=user, records=records, errors=errors)
    return load_detail(load_id, user)


@router.post("/cargas/manual")
def manual_load(payload: ManualChange, user: dict = Depends(require_role("CARGADOR"))):
    refs = references()
    raw = {**payload.datos, "Accion": payload.accion}
    normalized, errors = normalize_record(payload.tipo_registro, raw, refs)
    data = normalized["data"]
    if payload.tipo_registro == "EQUIVALENCIA":
        for field in ("Codigo_Desembolso", "Codigo_Consolidacion"):
            if data.get(field) and data[field] not in refs["products"]: errors.append((field, "PRODUCTO_INEXISTENTE", f"El producto {data[field]} no existe."))
    elif payload.tipo_registro == "CONDICION":
        for field in ("Codigo_Producto", "Codigo_Desembolso", "Codigo_Consolidacion"):
            if data.get(field) and data[field] not in refs["products"]: errors.append((field, "PRODUCTO_INEXISTENTE", f"El producto {data[field]} no existe."))
        if data.get("Modalidad_Convenio") and (data.get("Reglamentacion"), data["Modalidad_Convenio"]) not in refs["modalities"]: errors.append(("Modalidad_Convenio", "MODALIDAD_INVALIDA", "La modalidad no corresponde a la reglamentación."))
    elif payload.tipo_registro == "ORGANISMO_TASA":
        if data.get("Codigo_Organismo") not in refs["organisms"]: errors.append(("Codigo_Organismo", "ORGANISMO_INEXISTENTE", "El organismo no existe."))
        if data.get("Tasa_Pct") is not None and ("802_01", f"TASA_{int(data['Tasa_Pct'] * 100):03d}") not in refs["modalities"]: errors.append(("Tasa_Pct", "TASA_SIN_CUADRO", "No existe un cuadro/modalidad 802_01 para esta tasa."))
    record = {"sheet": "Cambio manual", "row": 1, "type": payload.tipo_registro, "normalized": normalized, "errors": errors}
    load_id = persist_load("MANUAL", None, None, payload.observaciones, user, [record], [])
    return load_detail(load_id, user)


@router.get("/cargas")
def loads(limit: int = 30, user: dict = Depends(require_role("CARGADOR"))):
    limit = max(1, min(limit, 100))
    return fetch_all(f"""SELECT TOP ({limit}) Carga_ID AS carga_id,Tipo_Carga AS tipo_carga,Nombre_Archivo AS nombre_archivo,Estado AS estado,
        Usuario_Carga AS usuario_carga,Fecha_Carga AS fecha_carga,Usuario_Aprobacion AS usuario_aprobacion,Fecha_Aprobacion AS fecha_aprobacion,
        Usuario_Publicacion AS usuario_publicacion,Fecha_Publicacion AS fecha_publicacion,Total_Registros AS total_registros,
        Total_Altas AS total_altas,Total_Modificaciones AS total_modificaciones,Total_Bajas AS total_bajas,Total_Errores AS total_errores
        FROM dbo.Cargas_Normativas ORDER BY Carga_ID DESC""")


@router.get("/cargas/{load_id}")
def load_detail(load_id: int, user: dict = Depends(require_role("CARGADOR"))):
    header = fetch_one("""SELECT Carga_ID AS carga_id,Tipo_Carga AS tipo_carga,Nombre_Archivo AS nombre_archivo,Estado AS estado,Usuario_Carga AS usuario_carga,
        Fecha_Carga AS fecha_carga,Usuario_Aprobacion AS usuario_aprobacion,Fecha_Aprobacion AS fecha_aprobacion,Usuario_Publicacion AS usuario_publicacion,
        Fecha_Publicacion AS fecha_publicacion,Observaciones AS observaciones,Total_Registros AS total_registros,Total_Altas AS total_altas,
        Total_Modificaciones AS total_modificaciones,Total_Bajas AS total_bajas,Total_Errores AS total_errores
        FROM dbo.Cargas_Normativas WHERE Carga_ID=?""", (load_id,))
    if not header: raise HTTPException(status_code=404, detail="La carga no existe")
    header["detalles"] = fetch_all("""SELECT Carga_Detalle_ID AS detalle_id,Hoja AS hoja,Numero_Fila AS fila,Tipo_Registro AS tipo_registro,
        Clave_Registro AS clave,Accion_Solicitada AS accion_solicitada,Accion_Detectada AS accion_detectada,Datos_JSON AS datos_json,
        Datos_Anteriores_JSON AS datos_anteriores_json,Estado AS estado FROM dbo.Carga_Normativa_Detalles WHERE Carga_ID=? ORDER BY Hoja,Numero_Fila""", (load_id,))
    header["errores"] = fetch_all("""SELECT Hoja AS hoja,Numero_Fila AS fila,Campo AS campo,Severidad AS severidad,Codigo_Error AS codigo,Mensaje AS mensaje
        FROM dbo.Carga_Normativa_Errores WHERE Carga_ID=? ORDER BY CASE Severidad WHEN 'ERROR' THEN 0 ELSE 1 END,Hoja,Numero_Fila""", (load_id,))

    allow_self_approval = os.getenv("ALLOW_SELF_APPROVAL", "false").lower() == "true"
    header["puede_aprobar"] = (
    header["estado"] == "VALIDADA"
    and ROLE_LEVEL[user["rol"]] >= ROLE_LEVEL["APROBADOR"]
    and (
        allow_self_approval
        or header["usuario_carga"].upper() != user["usuario"].upper()
    )
    )
    
    header["puede_publicar"] = header["estado"] == "APROBADA" and ROLE_LEVEL[user["rol"]] >= ROLE_LEVEL["APROBADOR"]
    return header


@router.post("/cargas/{load_id}/aprobar")
def approve(load_id: int, payload: Decision, user: dict = Depends(require_role("APROBADOR"))):
    load = fetch_one("SELECT Estado AS estado,Usuario_Carga AS usuario_carga FROM dbo.Cargas_Normativas WHERE Carga_ID=?", (load_id,))
    if not load: raise HTTPException(status_code=404, detail="La carga no existe")
    if load["estado"] != "VALIDADA": raise HTTPException(status_code=409, detail="Solo pueden aprobarse cargas validadas")
    if load["usuario_carga"].upper() == user["usuario"].upper() and os.getenv("ALLOW_SELF_APPROVAL", "false").lower() != "true":
        raise HTTPException(status_code=409, detail="Por doble control, quien cargó el cambio no puede aprobarlo")
    with get_connection() as connection:
        cursor = connection.cursor()
        cursor.execute("UPDATE dbo.Cargas_Normativas SET Estado='APROBADA',Usuario_Aprobacion=?,Fecha_Aprobacion=sysdatetime(),Observaciones=COALESCE(?,Observaciones) WHERE Carga_ID=? AND Estado='VALIDADA'", (user["usuario"], payload.observaciones, load_id))
        connection.commit()
    return load_detail(load_id, user)


@router.post("/cargas/{load_id}/rechazar")
def reject(load_id: int, payload: Decision, user: dict = Depends(require_role("APROBADOR"))):
    if not payload.observaciones: raise HTTPException(status_code=422, detail="Debe indicar el motivo del rechazo")
    with get_connection() as connection:
        cursor = connection.cursor()
        cursor.execute("UPDATE dbo.Cargas_Normativas SET Estado='RECHAZADA',Usuario_Aprobacion=?,Fecha_Aprobacion=sysdatetime(),Observaciones=? WHERE Carga_ID=? AND Estado IN ('VALIDADA','APROBADA')", (user["usuario"], payload.observaciones, load_id))
        if cursor.rowcount != 1: raise HTTPException(status_code=409, detail="La carga no está disponible para rechazo")
        connection.commit()
    return {"ok": True}


def ids(cursor, table, code_column, id_column, code):
    cursor.execute(f"SELECT {id_column} FROM dbo.{table} WHERE {code_column}=?", (code,))
    row = cursor.fetchone()
    if not row: raise ValueError(f"No existe {table}: {code}")
    return row[0]


def publish_record(cursor, load_id: int, detail_id: int, record_type: str, requested: str, data: dict):
    if record_type == "PRODUCTO":
        cursor.execute("SELECT Producto_ID FROM dbo.Productos WHERE Codigo=?", (data["Codigo"],)); row = cursor.fetchone()
        if requested == "BAJA": cursor.execute("UPDATE dbo.Productos SET Activo=0 WHERE Codigo=?", (data["Codigo"],))
        elif row: cursor.execute("UPDATE dbo.Productos SET Activo=? WHERE Codigo=?", (1 if data["Activo"] is not False else 0, data["Codigo"]))
        else: cursor.execute("INSERT dbo.Productos (Codigo,Activo) VALUES (?,?)", (data["Codigo"], 1 if data["Activo"] is not False else 0))
    elif record_type == "ORGANISMO":
        cursor.execute("SELECT Organismo_ID FROM dbo.Organismos WHERE Codigo=?", (data["Codigo"],)); row = cursor.fetchone()
        if requested == "BAJA": cursor.execute("UPDATE dbo.Organismos SET Activo=0 WHERE Codigo=?", (data["Codigo"],))
        elif row: cursor.execute("UPDATE dbo.Organismos SET Nombre=?,CUIT=?,Activo=? WHERE Codigo=?", (data["Nombre"], data["CUIT"], 1 if data["Activo"] is not False else 0, data["Codigo"]))
        else: cursor.execute("INSERT dbo.Organismos (Codigo,Nombre,CUIT,Activo) VALUES (?,?,?,?)", (data["Codigo"], data["Nombre"], data["CUIT"], 1 if data["Activo"] is not False else 0))
    elif record_type == "EQUIVALENCIA":
        desemb = ids(cursor,"Productos","Codigo","Producto_ID",data["Codigo_Desembolso"]); reg = ids(cursor,"Reglamentaciones","Codigo","Reglamentacion_ID",data["Reglamentacion"])
        if requested == "BAJA":
            cursor.execute("UPDATE dbo.Producto_Equivalencias SET Activa=0,Fecha_Hasta=COALESCE(?,Fecha_Hasta) WHERE Producto_Desembolso_ID=? AND Reglamentacion_ID=? AND Activa=1", (data["Fecha_Hasta"], desemb, reg))
        else:
            consol = ids(cursor,"Productos","Codigo","Producto_ID",data["Codigo_Consolidacion"])
            cursor.execute("UPDATE dbo.Producto_Equivalencias SET Fecha_Hasta=DATEADD(day,-1,?) WHERE Producto_Desembolso_ID=? AND Reglamentacion_ID=? AND Activa=1 AND Fecha_Desde<? AND (Fecha_Hasta IS NULL OR Fecha_Hasta>=?)", (data["Fecha_Desde"], desemb, reg, data["Fecha_Desde"], data["Fecha_Desde"]))
            cursor.execute("SELECT Producto_Equivalencia_ID FROM dbo.Producto_Equivalencias WHERE Producto_Desembolso_ID=? AND Reglamentacion_ID=? AND Fecha_Desde=?", (desemb, reg, data["Fecha_Desde"])); existing=cursor.fetchone()
            if existing: cursor.execute("UPDATE dbo.Producto_Equivalencias SET Producto_Consolidacion_ID=?,Fecha_Hasta=?,Activa=1,Carga_ID=? WHERE Producto_Equivalencia_ID=?", (consol,data["Fecha_Hasta"],load_id,existing[0]))
            else: cursor.execute("INSERT dbo.Producto_Equivalencias (Producto_Desembolso_ID,Producto_Consolidacion_ID,Reglamentacion_ID,Fecha_Desde,Fecha_Hasta,Activa,Carga_ID) VALUES (?,?,?,?,?,1,?)", (desemb,consol,reg,data["Fecha_Desde"],data["Fecha_Hasta"],load_id))
    elif record_type == "ORGANISMO_TASA":
        organism=ids(cursor,"Organismos","Codigo","Organismo_ID",data["Codigo_Organismo"]); reg=ids(cursor,"Reglamentaciones","Codigo","Reglamentacion_ID","802_01")
        if requested == "BAJA": cursor.execute("UPDATE dbo.Organismo_Condiciones_802_01 SET Vigencia_Hasta=?,Activa=1 WHERE Organismo_ID=? AND Tipo_Vivienda=? AND Activa=1", (data["Vigencia_Hasta"],organism,data["Tipo_Vivienda"]))
        else:
            modality_code=f"TASA_{int(Decimal(str(data['Tasa_Pct']))*100):03d}"
            cursor.execute("SELECT Modalidad_Convenio_ID FROM dbo.Modalidades_Convenio WHERE Reglamentacion_ID=? AND Codigo=?",(reg,modality_code)); modality=cursor.fetchone()
            if not modality: raise ValueError(f"No existe una modalidad 802_01 para la tasa {data['Tasa_Pct']}%")
            cursor.execute("UPDATE dbo.Organismo_Condiciones_802_01 SET Vigencia_Hasta=DATEADD(day,-1,?) WHERE Organismo_ID=? AND Tipo_Vivienda=? AND Activa=1 AND Vigencia_Desde<? AND Vigencia_Hasta>=?",(data["Vigencia_Desde"],organism,data["Tipo_Vivienda"],data["Vigencia_Desde"],data["Vigencia_Desde"]))
            cursor.execute("SELECT Organismo_Condicion_ID FROM dbo.Organismo_Condiciones_802_01 WHERE Organismo_ID=? AND Tipo_Vivienda=? AND Vigencia_Desde=?",(organism,data["Tipo_Vivienda"],data["Vigencia_Desde"])); existing=cursor.fetchone()
            values=(modality[0],data["Grupo_Pauta"],data["Vigencia_Hasta"],data["Adicional_Topeo_Pct"],1 if data["Aplica_Circular_3214"] else 0,data["Observaciones"],load_id)
            if existing: cursor.execute("UPDATE dbo.Organismo_Condiciones_802_01 SET Modalidad_Convenio_ID=?,Grupo_Pauta=?,Vigencia_Hasta=?,Adicional_Topeo_Pct=?,Aplica_Circular_3214=?,Observaciones=?,Fuente=N'Mantenimiento web',Activa=1,Carga_ID=? WHERE Organismo_Condicion_ID=?", values+(existing[0],))
            else: cursor.execute("INSERT dbo.Organismo_Condiciones_802_01 (Organismo_ID,Tipo_Vivienda,Modalidad_Convenio_ID,Grupo_Pauta,Vigencia_Desde,Vigencia_Hasta,Adicional_Topeo_Pct,Aplica_Circular_3214,Observaciones,Fuente,Activa,Carga_ID) VALUES (?,?,?,?,?,?,?,?,?,N'Mantenimiento web',1,?)",(organism,data["Tipo_Vivienda"],modality[0],data["Grupo_Pauta"],data["Vigencia_Desde"],data["Vigencia_Hasta"],data["Adicional_Topeo_Pct"],1 if data["Aplica_Circular_3214"] else 0,data["Observaciones"],load_id))
    elif record_type == "CONDICION":
        cursor.execute("SELECT TOP(1) Condicion_ID,Fecha_Desde FROM dbo.Condiciones_Reglamentacion WHERE Clave_Mantenimiento=? AND Activa=1 ORDER BY ISNULL(Fecha_Desde,'19000101') DESC,Condicion_ID DESC",(data["Clave_Mantenimiento"],)); previous=cursor.fetchone()
        if requested == "BAJA":
            if not previous: raise ValueError("No existe una condición vigente para dar de baja")
            cursor.execute("UPDATE dbo.Condiciones_Reglamentacion SET Fecha_Hasta=?,Activa=1 WHERE Condicion_ID=?",(data["Fecha_Hasta"] or date.today(),previous[0]))
            return
        reg=ids(cursor,"Reglamentaciones","Codigo","Reglamentacion_ID",data["Reglamentacion"]); area=ids(cursor,"Ambitos_Geograficos","Codigo","Ambito_Geografico_ID",data["Ambito_Geografico"]); dest=ids(cursor,"Destinos","Codigo","Destino_ID",data["Destino"]); user_type=ids(cursor,"Tipos_Usuario","Codigo","Tipo_Usuario_ID",data["Tipo_Usuario"])
        modality=None
        if data["Modalidad_Convenio"]:
            cursor.execute("SELECT Modalidad_Convenio_ID FROM dbo.Modalidades_Convenio WHERE Reglamentacion_ID=? AND Codigo=?",(reg,data["Modalidad_Convenio"])); row=cursor.fetchone(); modality=row[0] if row else None
            if modality is None: raise ValueError("La modalidad no corresponde a la reglamentación")
        if previous and data["Fecha_Desde"] and (previous[1] is None or data["Fecha_Desde"]>previous[1]):
            cursor.execute("UPDATE dbo.Condiciones_Reglamentacion SET Fecha_Hasta=DATEADD(day,-1,?),Activa=1 WHERE Condicion_ID=?",(data["Fecha_Desde"],previous[0])); previous=None
        code=f"M{load_id:010d}{detail_id%1000000:06d}"
        target_id=previous[0] if previous else None
        values=(reg,area,dest,user_type,modality,data["Fecha_Desde"],data["Fecha_Hasta"],data["UVA_Desde"],data["UVA_Desde_Inclusive"],data["UVA_Hasta"],data["UVA_Hasta_Inclusive"],data["Tasa_Aplicable_Pct"],data["Permite_Topeo"],data["Cobra_Prima_Topeo"],data["Referencia_Fuente"],data["Observaciones"],data["Clave_Mantenimiento"],load_id)
        if target_id: cursor.execute("""UPDATE dbo.Condiciones_Reglamentacion SET Reglamentacion_ID=?,Ambito_Geografico_ID=?,Destino_ID=?,Tipo_Usuario_ID=?,Modalidad_Convenio_ID=?,Fecha_Desde=?,Fecha_Hasta=?,UVA_Desde=?,UVA_Desde_Inclusive=?,UVA_Hasta=?,UVA_Hasta_Inclusive=?,Tasa_Aplicable_Pct=?,Permite_Topeo=?,Cobra_Prima_Topeo=?,Referencia_Fuente=?,Observaciones=?,Clave_Mantenimiento=?,Carga_ID=?,Activa=1 WHERE Condicion_ID=?""",values+(target_id,)); condition_id=target_id
        else:
            cursor.execute("""INSERT dbo.Condiciones_Reglamentacion (Codigo_Condicion,Reglamentacion_ID,Ambito_Geografico_ID,Destino_ID,Tipo_Usuario_ID,Modalidad_Convenio_ID,Fecha_Desde,Fecha_Hasta,UVA_Desde,UVA_Desde_Inclusive,UVA_Hasta,UVA_Hasta_Inclusive,Tasa_Aplicable_Pct,Permite_Topeo,Cobra_Prima_Topeo,Referencia_Fuente,Observaciones,Clave_Mantenimiento,Carga_ID,Activa) OUTPUT INSERTED.Condicion_ID VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)""",(code,)+values); condition_id=cursor.fetchone()[0]
        cursor.execute("DELETE dbo.Condicion_Productos WHERE Condicion_ID=?",(condition_id,))
        for role,field in (("UNICO","Codigo_Producto"),("DESEMBOLSO","Codigo_Desembolso"),("CONSOLIDACION","Codigo_Consolidacion")):
            if data[field]: cursor.execute("INSERT dbo.Condicion_Productos (Condicion_ID,Producto_ID,Rol_Producto) SELECT ?,Producto_ID,? FROM dbo.Productos WHERE Codigo=?",(condition_id,role,data[field]))


@router.post("/cargas/{load_id}/publicar")
def publish(load_id: int, user: dict = Depends(require_role("APROBADOR"))):
    load = fetch_one("SELECT Estado AS estado FROM dbo.Cargas_Normativas WHERE Carga_ID=?",(load_id,))
    if not load: raise HTTPException(status_code=404,detail="La carga no existe")
    if load["estado"]!="APROBADA": raise HTTPException(status_code=409,detail="La carga debe estar aprobada antes de publicar")
    details=fetch_all("SELECT Carga_Detalle_ID AS id,Tipo_Registro AS tipo,Accion_Solicitada AS accion,Accion_Detectada AS detectada,Datos_JSON AS datos,Datos_Anteriores_JSON AS anteriores FROM dbo.Carga_Normativa_Detalles WHERE Carga_ID=? AND Estado='VALIDO' ORDER BY CASE Tipo_Registro WHEN 'PRODUCTO' THEN 1 WHEN 'ORGANISMO' THEN 2 WHEN 'EQUIVALENCIA' THEN 3 WHEN 'CONDICION' THEN 4 ELSE 5 END,Carga_Detalle_ID",(load_id,))
    try:
        with get_connection() as connection:
            cursor=connection.cursor()
            cursor.execute("UPDATE dbo.Cargas_Normativas SET Estado='PUBLICADA',Usuario_Publicacion=?,Fecha_Publicacion=sysdatetime() WHERE Carga_ID=? AND Estado='APROBADA'",(user["usuario"],load_id))
            if cursor.rowcount!=1: raise ValueError("La carga cambió de estado")
            for item in details:
                data=json.loads(item["datos"])
                for field,value in list(data.items()):
                    if field.startswith("Fecha_") or field.startswith("Vigencia_"):
                        data[field]=date.fromisoformat(value) if value else None
                if item["detectada"]!="SIN_CAMBIOS": publish_record(cursor,load_id,item["id"],item["tipo"],item["accion"],data)
                cursor.execute("INSERT dbo.Auditoria_Cambios_Normativos (Carga_ID,Tipo_Registro,Clave_Registro,Accion,Datos_Anteriores_JSON,Datos_Nuevos_JSON,Usuario_Windows) SELECT ?,Tipo_Registro,Clave_Registro,?,Datos_Anteriores_JSON,Datos_JSON,? FROM dbo.Carga_Normativa_Detalles WHERE Carga_Detalle_ID=?",(load_id,item["detectada"] if item["detectada"]!="SIN_CAMBIOS" else "MODIFICACION",user["usuario"],item["id"]))
                cursor.execute("UPDATE dbo.Carga_Normativa_Detalles SET Estado='PUBLICADO' WHERE Carga_Detalle_ID=?",(item["id"],))
            connection.commit()
    except Exception as exc:
        raise HTTPException(status_code=409,detail=f"No se publicó ningún cambio. La transacción fue revertida: {exc}") from exc
    return load_detail(load_id,user)


@router.get("/auditoria")
def audit(limit: int=100, _: dict=Depends(require_role("APROBADOR"))):
    limit=max(1,min(limit,500))
    return fetch_all(f"SELECT TOP ({limit}) Auditoria_ID AS auditoria_id,Carga_ID AS carga_id,Tipo_Registro AS tipo_registro,Clave_Registro AS clave,Accion AS accion,Usuario_Windows AS usuario,Fecha_Cambio AS fecha_cambio FROM dbo.Auditoria_Cambios_Normativos ORDER BY Auditoria_ID DESC")


@router.get("/usuarios")
def users(_: dict=Depends(require_role("ADMINISTRADOR"))):
    return fetch_all("SELECT Usuario_Windows AS usuario,Nombre_Mostrar AS nombre,Rol AS rol,Activo AS activo,Fecha_Alta AS fecha_alta FROM dbo.Usuarios_Aplicacion ORDER BY Usuario_Windows")


class UserUpdate(BaseModel):
    usuario: str=Field(min_length=2,max_length=150)
    nombre: str|None=Field(default=None,max_length=180)
    rol: str=Field(pattern=r"^(CONSULTOR|CARGADOR|APROBADOR|ADMINISTRADOR)$")
    activo: bool=True


@router.post("/usuarios")
def save_user(payload: UserUpdate, _: dict=Depends(require_role("ADMINISTRADOR"))):
    with get_connection() as connection:
        cursor=connection.cursor(); cursor.execute("SELECT 1 FROM dbo.Usuarios_Aplicacion WHERE Usuario_Windows=?",(payload.usuario,))
        if cursor.fetchone(): cursor.execute("UPDATE dbo.Usuarios_Aplicacion SET Nombre_Mostrar=?,Rol=?,Activo=? WHERE Usuario_Windows=?",(payload.nombre,payload.rol,payload.activo,payload.usuario))
        else: cursor.execute("INSERT dbo.Usuarios_Aplicacion (Usuario_Windows,Nombre_Mostrar,Rol,Activo) VALUES (?,?,?,?)",(payload.usuario,payload.nombre,payload.rol,payload.activo))
        connection.commit()
    return {"ok":True}
